import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_all_vacancies_for_export

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

def format_sheet_header(ws, headers: list):
    """Стилизует строку заголовков и закрепляет верхнюю панель."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='medium', color='1F4E79')
    )
    
    ws.append(headers)
    ws.row_dimensions[1].height = 28
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws.freeze_panes = "A2"

def autofit_columns(ws, max_cols: int):
    """Автоматически подбирает ширину колонок с учетом содержимого."""
    for col in range(1, max_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

def fill_sheet_data(ws, vacancies: list):
    """Заполняет лист вакансиями со стилями, границами и ссылками."""
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    status_fills = {
        'Оффер': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        'Собеседование': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
        'Тестовое задание': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        'Откликнулся': PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
        'Отказ': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    }
    
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    regular_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    
    for idx, v in enumerate(vacancies, start=1):
        row_num = idx + 1
        is_zebra = (idx % 2 == 0)
        current_fill = zebra_fill if is_zebra else white_fill
        
        raw_date = v.get('created_at', '') or v.get('published_at', '')
        date_str = raw_date[:16].replace('T', ' ') if raw_date else datetime.now().strftime('%Y-%m-%d %H:%M')
        status = v.get('status', 'Не откликался') or "Не откликался"
        
        row_data = [
            idx,
            v.get('title', 'Без названия'),
            v.get('company', 'Компания не указана'),
            v.get('salary', 'По договоренности'),
            v.get('source', 'Сайт'),
            "🔗 Открыть вакансию",
            date_str,
            status,
            ""
        ]
        
        ws.append(row_data)
        ws.row_dimensions[row_num].height = 22
        
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = current_fill
            cell.border = thin_border
            cell.font = regular_font
            
            if col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [2, 3]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_num == 4:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if "от" in str(cell.value).lower() or "₽" in str(cell.value) or "$" in str(cell.value):
                    cell.font = bold_font
            elif col_num == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 6:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                target_url = v.get('url', '')
                if target_url:
                    cell.hyperlink = target_url
                    cell.font = link_font
            elif col_num == 7:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if status in status_fills:
                    cell.fill = status_fills[status]
                    cell.font = bold_font
            elif col_num == 9:
                cell.alignment = Alignment(horizontal="left", vertical="center")

def create_excel_export(output_filename: str = "vacancies_export.xlsx") -> str:
    all_vacancies = get_all_vacancies_for_export()
    
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "📋 Все вакансии"
    
    headers = [
        "№", 
        "Должность", 
        "Компания", 
        "Зарплата", 
        "Источник", 
        "Прямая ссылка", 
        "Дата добавления", 
        "Статус отклика", 
        "Мои заметки"
    ]
    
    # 1. Лист «Все вакансии»
    format_sheet_header(ws_all, headers)
    fill_sheet_data(ws_all, all_vacancies)
    autofit_columns(ws_all, len(headers))
    if len(all_vacancies) > 0:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_vacancies) + 1}"
        
    # 2. Лист «HeadHunter»
    hh_vacancies = [v for v in all_vacancies if v.get('source') == 'HeadHunter']
    if hh_vacancies:
        ws_hh = wb.create_sheet(title="🔴 HeadHunter")
        format_sheet_header(ws_hh, headers)
        fill_sheet_data(ws_hh, hh_vacancies)
        autofit_columns(ws_hh, len(headers))
        ws_hh.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(hh_vacancies) + 1}"
        
    # 3. Лист «Хабр Карьера»
    habr_vacancies = [v for v in all_vacancies if v.get('source') == 'Хабр Карьера']
    if habr_vacancies:
        ws_habr = wb.create_sheet(title="🔵 Хабр Карьера")
        format_sheet_header(ws_habr, headers)
        fill_sheet_data(ws_habr, habr_vacancies)
        autofit_columns(ws_habr, len(headers))
        ws_habr.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(habr_vacancies) + 1}"
        
    # 4. Лист «Telegram стажировки»
    tg_vacancies = [v for v in all_vacancies if v.get('source') == 'Telegram']
    if tg_vacancies:
        ws_tg = wb.create_sheet(title="📱 Telegram")
        format_sheet_header(ws_tg, headers)
        fill_sheet_data(ws_tg, tg_vacancies)
        autofit_columns(ws_tg, len(headers))
        ws_tg.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(tg_vacancies) + 1}"

    # 5. Лист «Авито Работа»
    avito_vacancies = [v for v in all_vacancies if v.get('source') == 'Авито Работа']
    if avito_vacancies:
        ws_avito = wb.create_sheet(title="🟡 Авито")
        format_sheet_header(ws_avito, headers)
        fill_sheet_data(ws_avito, avito_vacancies)
        autofit_columns(ws_avito, len(headers))
        ws_avito.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(avito_vacancies) + 1}"

    output_path = os.path.join(os.path.dirname(__file__), output_filename)
    wb.save(output_path)
    return output_path

if __name__ == '__main__':
    path = create_excel_export()
    print("Экспорт успешно создан:", path)
